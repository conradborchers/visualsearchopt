import glob
import os
import openpyxl
import fileinput
import re

# Original recoding
d_recode = {
# ditionary holding all participant code fixes
}

d_recode_rev = {value: key for key, value in d_recode_rev.items()}

all_codes = {
# ditionary holding all participant codes
}

d_anon_crosswalk = dict()

# Strategy:
# For all codes: See if they have a lookup/correction
# If not: Assign new anon ID and create entry in anon crosswalk
# If yes: Assign anon ID of lookup (create entry if there is none)

count = 1
for code in all_codes.keys():
  if code not in d_recode:
    if code not in d_anon_crosswalk:
      d_anon_crosswalk[code] = 'anon'+str(count)
      count+=1
    else:
      continue
  else:
    code_recoded = d_recode[code]
    if code_recoded in d_anon_crosswalk:
      d_anon_crosswalk[code] = d_anon_crosswalk[code_recoded]
    else:
      d_anon_crosswalk[code_recoded] = 'anon'+str(count)
      d_anon_crosswalk[code] = d_anon_crosswalk[code_recoded]
      count+=1

# Create new anon files (also functions.R)
os.chdir("data/")

# Excel files
for file in glob.glob("*.xlsx"):
  
    outfile = file.replace('.xlsx', '_ANON.xlsx')
    outfile = '../data_anon/' + outfile
    
    wb = openpyxl.load_workbook(file)
    ws = wb["Tabelle1"]
    
    i=0
    for code in d_anon_crosswalk.keys():
        for r in range(1,ws.max_row+1):
            for c in range(1,ws.max_column+1):
                s = ws.cell(r,c).value
                if s != None and str(code) in str(s): 
                    ws.cell(r,c).value = s.replace(code, d_anon_crosswalk[code])
                    i += 1
    
    wb.save(outfile)

# Text files
for filename in glob.glob("*.txt"):
    outfile = filename.replace('.txt', '_ANON.txt')
    outfile = '../data_anon/' + outfile
    with open(outfile, 'w') as outf:
        with fileinput.FileInput(filename, inplace=False, backup='.bak') as file:
            for line in file:
                for code in d_anon_crosswalk.keys():
                    rgx = re.compile(code)
                    line = rgx.sub(d_anon_crosswalk[code], line)
                outf.write(line)

# CSV files
for filename in glob.glob("*.csv"):
    outfile = filename.replace('.csv', '_ANON.csv')
    outfile = '../data_anon/' + outfile
    with open(outfile, 'w') as outf:
        with fileinput.FileInput(filename, inplace=False, backup='.bak') as file:
            for line in file:
                for code in d_anon_crosswalk.keys():
                    rgx = re.compile(code)
                    line = rgx.sub(d_anon_crosswalk[code], line)
                outf.write(line)

# R File
os.chdir("../R")
filename = 'functions.R'
outfile = 'functions_anon.R'
with open(outfile, 'w') as outf:
    with fileinput.FileInput(filename, inplace=False, backup='.bak') as file:
        for line in file:
            for code in d_anon_crosswalk.keys():
                rgx = re.compile(code)
                line = rgx.sub(d_anon_crosswalk[code], line)
            outf.write(line)
